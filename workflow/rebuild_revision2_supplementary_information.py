#!/usr/bin/env python3
"""Rebuild revision2 SupplementaryInformation.xlsx from ARG-rooted ZI results.

The script copies the original supplementary workbook and replaces the tables
whose values depend on the ARG-rooted SFRatios analysis:
Tables 2-8 and 10-12. Tables 3-5 require bootstrap confidence intervals.
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from collections import OrderedDict
from statistics import mean
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_OLD_XLSX = Path(
    "/mnt/d/genemod/better_dNdS_models/popgen/Drosophila_SFS_and_SFRatios/"
    "codon2NS_manuscript/MBE/revision/SupplementaryInformation.xlsx"
)
DEFAULT_OUT_XLSX = Path(
    "/mnt/d/genemod/better_dNdS_models/popgen/Drosophila_SFS_and_SFRatios/"
    "codon2NS_manuscript/MBE/revision2/SupplementaryInformation_revision2.xlsx"
)
DEFAULT_MAIN_WORK = Path(
    "/mnt/d/genemod/better_dNdS_models/popgen/SFRatios_pipeline_7_9_2026/"
    "ZIResults/singer10rooted_SFRatios_n160"
)
DEFAULT_ESTSFS_P0999_WORK = Path(
    "/mnt/d/genemod/better_dNdS_models/popgen/SFRatios_pipeline_7_9_2026/"
    "ZIResults/est_sfs_5outgroups_rootp0.999/SFRatioswork"
)
DEFAULT_IMPUTED_WORK = Path(
    "/mnt/d/genemod/better_dNdS_models/drosophila/species_2Ns_comparison/"
    "DmelZI/argwork/singer10rooted_SFRatios_work"
)
DEFAULT_DECONV_WORK = Path(
    "/mnt/d/genemod/better_dNdS_models/popgen/SFRatios_pipeline_7_9_2026/"
    "ZIResults/singer10rooted_SFRatios_n160_deconv0.03"
)
DEFAULT_BOOT_CI = Path(
    "/mnt/d/genemod/better_dNdS_models/popgen/Drosophila_SFS_and_SFRatios/"
    "codon2NS_manuscript/MBE/revision2/bootstrap120_n160_singer10rooted/checkbootstrap_120_datasets.txt"
)
DEFAULT_BUGFIX_DIR = Path(
    "/mnt/d/genemod/better_dNdS_models/popgen/Drosophila_SFS_and_SFRatios/"
    "codon2NS_manuscript/MBE/revision2"
)

SHEET2 = "Table 2. Codon Pair SFSs "
SHEET3 = "Table 3. Initial 2Ns estimates"
SHEET4 = "Table 4. Fitted_2Ns_estimates"
SHEET5 = "Table 5. g estimates"
SHEET6 = "Table 6. g pipeline variants"
SHEET7 = "Table 7 Amino Acid properites"
SHEET8 = "Table 8 Codon Fitness Model Fit"
SHEET10 = "Table 10, codon slopes intercep"
SHEET11 = "Table 11. codon slopes and 2Ns"
SHEET12 = "Table 12. expression-scaled mod"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create the revision2 supplementary workbook by replacing SFS, "
            "2Ns, fitted 2Ns, g, and pipeline-variant tables."
        )
    )
    parser.add_argument(
        "-i",
        "--input-xlsx",
        type=Path,
        default=DEFAULT_OLD_XLSX,
        help=f"Original supplementary workbook to copy. Default: {DEFAULT_OLD_XLSX}",
    )
    parser.add_argument(
        "-o",
        "--output-xlsx",
        type=Path,
        default=DEFAULT_OUT_XLSX,
        help=f"Output revision2 workbook. Default: {DEFAULT_OUT_XLSX}",
    )
    parser.add_argument(
        "-m",
        "--main-work",
        type=Path,
        default=DEFAULT_MAIN_WORK,
        help=f"Main ARG-rooted SFRatios work directory. Default: {DEFAULT_MAIN_WORK}",
    )
    parser.add_argument(
        "-p",
        "--prefix",
        default="ZIResults_singer10rooted_n160",
        help="Main result filename prefix. Default: ZIResults_singer10rooted_n160",
    )
    parser.add_argument(
        "--estsfs-p0999-work",
        type=Path,
        default=DEFAULT_ESTSFS_P0999_WORK,
        help=f"est-sfs p>=0.999 comparison work directory. Default: {DEFAULT_ESTSFS_P0999_WORK}",
    )
    parser.add_argument(
        "-b",
        "--bootstrap-ci",
        type=Path,
        default=DEFAULT_BOOT_CI,
        help=f"Bootstrap CI summary from 120 resamples. Default: {DEFAULT_BOOT_CI}",
    )
    parser.add_argument(
        "--imputed-work",
        type=Path,
        default=DEFAULT_IMPUTED_WORK,
        help=f"Imputed n=190 SFRatios work directory. Default: {DEFAULT_IMPUTED_WORK}",
    )
    parser.add_argument(
        "--deconv-work",
        type=Path,
        default=DEFAULT_DECONV_WORK,
        help=f"3%% deconvolution SFRatios work directory. Default: {DEFAULT_DECONV_WORK}",
    )
    parser.add_argument(
        "--allow-missing-ci",
        action="store_true",
        help=(
            "Write point estimates and leave CI columns blank if the bootstrap "
            "summary is absent. Default: abort when CI file is missing."
        ),
    )
    parser.add_argument(
        "--bugfix-dir",
        type=Path,
        default=DEFAULT_BUGFIX_DIR,
        help=f"Directory containing rerun Figure 3/12 model outputs. Default: {DEFAULT_BUGFIX_DIR}",
    )
    return parser.parse_args()


def clear_sheet(ws) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def write_rows(ws, rows: list[list[object]]) -> None:
    clear_sheet(ws)
    for row in rows:
        ws.append(row)


def parse_sfs(path: Path) -> list[list[object]]:
    rows: list[list[object]] = []
    lines = [line.strip() for line in path.read_text().splitlines() if line.strip()]
    if len(lines) % 4 != 0:
        raise ValueError(f"{path} has {len(lines)} nonblank lines, expected a multiple of 4")
    for i in range(0, len(lines), 4):
        syn_label = lines[i].split()
        intron_label = lines[i + 2].split()
        if len(syn_label) != 2 or syn_label[0] != "Synonymous":
            raise ValueError(f"unexpected synonymous label at line {i + 1}: {lines[i]}")
        if len(intron_label) != 3 or intron_label[:2] != ["Intron", "for"]:
            raise ValueError(f"unexpected intron label at line {i + 3}: {lines[i + 2]}")
        rows.append(["Synonymous_", syn_label[1]])
        rows.append([int(x) for x in lines[i + 1].split()])
        rows.append([f"Intron_for_{intron_label[2]}"])
        rows.append([int(x) for x in lines[i + 3].split()])
    return rows


def codon_pair_from_filename(filename: str) -> str:
    match = re.search(r"Synonymous_([ACGT]{3})_([ACGT]{3})_Qratio", filename)
    if not match:
        raise ValueError(f"cannot parse codon pair from {filename}")
    return f"{match.group(1)}_{match.group(2)}"


def parse_sfratios_summary(path: Path) -> OrderedDict[str, float]:
    vals: OrderedDict[str, float] = OrderedDict()
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            vals[codon_pair_from_filename(row["filename"])] = float(row["2Ns"])
    return vals


def parse_ls_codon_values(path: Path) -> OrderedDict[str, float]:
    codons: OrderedDict[str, float] = OrderedDict()
    in_table = False
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if line == "AA\tCodon\t2Ns":
            in_table = True
            continue
        if not in_table or not line:
            continue
        parts = line.split()
        if len(parts) >= 3 and re.fullmatch(r"[ACGT]{3}", parts[1]):
            codons[parts[1]] = float(parts[2])
    if not codons:
        raise ValueError(f"no codon g values found in {path}")
    return codons


def parse_ls_model_fit(path: Path) -> list[list[object]]:
    rows: list[list[object]] = [["Model Fitting"], ["Correlations of estimated 2Ns and fitted 2Ns for each amino acid "], [], ["AA", "#codons", "Correlation*", "CorrSim_p_(prop._sim_higher)**"]]
    in_table = False
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if line.startswith("AA\t#codons\tbest_codon"):
            in_table = True
            continue
        if in_table and line == "AA\tCodon\t2Ns":
            break
        if not in_table or not line:
            continue
        parts = line.split("\t")
        if len(parts) >= 8 and re.fullmatch(r"[A-Z]", parts[0]):
            p_value = parts[7]
            rows.append([parts[0], int(parts[1]), float(parts[5]), p_value if p_value == "na" else float(p_value)])
    rows.extend([
        [],
        ["*for k=2 amino acids, fitted values equal estimated values minus the mean value, i.e. correlation must equal 1. "],
        [],
        ["** the proportion of randomly permuted matrics with higher correlation values. "],
    ])
    return rows


def fitted_pairs_from_codons(pair_order: list[str], codon_g: dict[str, float]) -> OrderedDict[str, float]:
    fitted: OrderedDict[str, float] = OrderedDict()
    for pair in pair_order:
        src, dst = pair.split("_")
        fitted[pair] = codon_g[dst] - codon_g[src]
    return fitted


def parse_bootstrap_ci(path: Path) -> tuple[dict[str, tuple[float, float]], dict[str, tuple[float, float]], dict[str, tuple[float, float]]]:
    """Parse bootstrap CI summary.

    Expected sections are produced by run_SFRatios_and_LeastSquares_on_bootstrap_samples.py
    and include primary2Ns, fitted2Ns, and estg values. The parser is deliberately
    permissive about headers so it can handle either tabular or whitespace output.
    """
    primary: dict[str, tuple[float, float]] = {}
    fitted: dict[str, tuple[float, float]] = {}
    estg: dict[str, tuple[float, float]] = {}
    section: str | None = None
    in_table = False
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line:
            in_table = False
            continue
        lower = line.lower()
        if lower == "primary 2ns 95% confidence intervals":
            section = "primary"
            in_table = False
            continue
        if lower == "fitted 2ns 95% confidence intervals":
            section = "fitted"
            in_table = False
            continue
        if lower == "estimated g 95% confidence intervals":
            section = "estg"
            in_table = False
            continue
        if section in {"primary", "fitted"} and line.startswith("CodonPair\tMean\tLowCI\tHiCI"):
            in_table = True
            continue
        if section == "estg" and line.startswith("Codon\tMean\tLowCI\tHiCI"):
            in_table = True
            continue
        if line.startswith("Mean width:"):
            in_table = False
            continue
        if not in_table:
            continue
        parts = re.split(r"\s+", line)
        if section in {"primary", "fitted"} and len(parts) >= 4 and re.fullmatch(r"[ACGT]{3}_[ACGT]{3}", parts[0]):
            lo, hi = float(parts[2]), float(parts[3])
            (primary if section == "primary" else fitted)[parts[0]] = (lo, hi)
        elif section == "estg" and len(parts) >= 4 and re.fullmatch(r"[ACGT]{3}", parts[0]):
            estg[parts[0]] = (float(parts[2]), float(parts[3]))
    return primary, fitted, estg


def rows_with_ci(
    header: list[str],
    estimates: OrderedDict[str, float],
    ci: dict[str, tuple[float, float]] | None,
) -> list[list[object]]:
    rows = [header]
    for key, est in estimates.items():
        if ci and key in ci:
            lo, hi = ci[key]
            rows.append([key, est, lo, hi, hi - lo])
        else:
            rows.append([key, est, None, None, None])
    return rows


def existing_first_column_order(ws, skip_rows: int = 1) -> list[str]:
    order: list[str] = []
    for row in ws.iter_rows(min_row=skip_rows + 1, max_col=1, values_only=True):
        value = row[0]
        if value is not None:
            order.append(str(value))
    return order


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def tsv_rows(path: Path) -> list[list[object]]:
    with path.open(newline="") as handle:
        return [row for row in csv.reader(handle, delimiter="\t")]


def parse_key_value_tsv(path: Path) -> dict[str, str]:
    rows = read_tsv(path)
    return {row["Statistic"]: row["Value"] for row in rows}


def numeric(value: str) -> object:
    if value == "":
        return None
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


def convert_table_values(rows: list[list[object]]) -> list[list[object]]:
    return [[numeric(v) if isinstance(v, str) else v for v in row] for row in rows]


def update_table7(ws, codon_g: dict[str, float]) -> None:
    values_by_aa: dict[str, list[float]] = {}
    genetic_code = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S",
        "TAT": "Y", "TAC": "Y", "TGT": "C", "TGC": "C", "CTT": "L", "CTC": "L", "CTA": "L", "CTG": "L",
        "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q", "CAG": "Q",
        "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "I", "ACT": "T",
        "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "R", "AGG": "R", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A",
        "GCC": "A", "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G",
        "GGC": "G", "GGA": "G", "GGG": "G",
    }
    for codon, val in codon_g.items():
        values_by_aa.setdefault(genetic_code[codon], []).append(abs(val))
    mean_abs = {aa: mean(vals) for aa, vals in values_by_aa.items()}
    for row in range(2, ws.max_row + 1):
        aa = ws.cell(row, 1).value
        if aa in mean_abs:
            ws.cell(row, 3).value = mean_abs[aa]


def build_table11_rows(comparison_path: Path, codon_slope_path: Path) -> list[list[object]]:
    stats = parse_key_value_tsv(comparison_path)
    rows: list[list[object]] = [
        ["Section: Summary statistics for codon-specific expression slopes versus codon fitness"],
        ["Statistic", "Value", "Notes"],
        ["N", numeric(stats["n"]), "Synonymous codons"],
        ["Pearson r", numeric(stats["pearson_r"])],
        ["Pearson p-value", numeric(stats["pearson_p"])],
        ["Spearman rho", numeric(stats["spearman_rho"])],
        ["Spearman p-value", numeric(stats["spearman_p"])],
        ["OLS slope", numeric(stats["ols_slope"])],
        ["OLS intercept", numeric(stats["ols_intercept"])],
        ["OLS R-squared", numeric(stats["ols_r_squared"])],
        ["OLS slope p-value", numeric(stats["ols_slope_p"])],
        ["SMA slope", numeric(stats["sma_slope"]), "Standardized major-axis regression"],
        ["SMA intercept", numeric(stats["sma_intercept"]), "Standardized major-axis regression"],
        ["Cluster-bootstrap SMA slope CI low", numeric(stats["sma_slope_ci_low_cluster_bootstrap"]), "95% confidence interval"],
        ["Cluster-bootstrap SMA slope CI high", numeric(stats["sma_slope_ci_high_cluster_bootstrap"]), "95% confidence interval"],
        [],
        ["Section: Codon-specific estimates used in the correlation and SMA analyses"],
    ]
    rows.extend(convert_table_values(tsv_rows(codon_slope_path)))
    return rows


def build_table12_rows(summary_path: Path, aggregate_path: Path) -> list[list[object]]:
    stats = parse_key_value_tsv(summary_path)
    rows: list[list[object]] = [
        ["Section: Expression-scaled mutation-selection-drift effect estimates"],
        ["Statistic", "Value", "Notes"],
        ["Model formula", stats["model_formula"], "Probabilities normalized within synonymous codon family"],
        ["Selection exponent multiplier", numeric(stats["selection_exponent_multiplier"]), "Uses the same selection term as the expected codon frequency model"],
        ["Expression transform", stats["expression_transform"], "Expression predictor used for the fitted selection scale"],
        ["Mean log expression", numeric(stats["log_expression_mean"]), "Used for standardizing log(expression + 1)"],
        ["SD log expression", numeric(stats["log_expression_sd"]), "Used for standardizing log(expression + 1)"],
        ["Expression zero count", numeric(stats["expression_zero_count"])],
        ["Constant vs mutation LRT", numeric(stats["constant_vs_mutation_LRT"])],
        ["Constant vs mutation df", numeric(stats["constant_vs_mutation_df"])],
        ["Constant vs mutation p-value", numeric(stats["constant_vs_mutation_p"])],
        ["Expression vs constant LRT", numeric(stats["expression_vs_constant_LRT"])],
        ["Expression vs constant df", numeric(stats["expression_vs_constant_df"])],
        ["Expression vs constant p-value", numeric(stats["expression_vs_constant_p"])],
        ["Expression-scaled alpha", numeric(stats["alpha"]), "Intercept for log selection scale"],
        ["Expression-scaled beta", numeric(stats["beta_standardized_log_expression"]), "Effect of standardized log(expression + 1)"],
        ["Lambda at mean log-expression", numeric(stats["lambda_at_mean_log_expression"]), "Selection scale for a gene with average standardized log expression"],
        ["Standardized predictor change for 10-fold expression increase", numeric(stats["tenfold_predictor_change"]), "log(10) / SD[log(expression + 1)]"],
        ["Selection-scale fold change per 10-fold expression increase", numeric(stats["selection_scale_fold_change_per_10fold_expression"])],
        ["Selection-scale fold-change CI low", numeric(stats["selection_scale_fold_change_per_10fold_expression_ci_low"]), "200-bootstrap 95% confidence interval"],
        ["Selection-scale fold-change CI high", numeric(stats["selection_scale_fold_change_per_10fold_expression_ci_high"]), "200-bootstrap 95% confidence interval"],
        ["Bootstrap resampling scheme", "Genes resampled with replacement; within each resampled gene and amino-acid family, synonymous codon counts resampled from the observed count vector conditional on N_ar", "Codon identities, amino-acid families, mutation frequencies, and fitness estimates were held fixed"],
        ["Successful bootstrap replicates", 200],
        [],
        ["Section: Codon-level observed and predicted aggregate frequencies used to summarize model fit"],
    ]
    rows.extend(convert_table_values(tsv_rows(aggregate_path)))
    return rows


def require(path: Path, label: str) -> Path:
    if not path.exists():
        raise FileNotFoundError(f"missing {label}: {path}")
    return path


def main() -> int:
    args = parse_args()
    input_xlsx = require(args.input_xlsx, "input workbook")
    main_sfs = require(args.main_work / f"{args.prefix}_SFSs.txt", "main SFS file")
    main_summary = require(args.main_work / f"{args.prefix}_SFRatios_summary.txt", "main SFRatios summary")
    main_ls = require(args.main_work / f"{args.prefix}_LeastSquares_analysis.txt", "main least-squares analysis")
    estsfs_p0999_ls = require(
        args.estsfs_p0999_work / "ZI_estSFSrootp0.999_LeastSquares_analysis.txt",
        "est-sfs p>=0.999 least-squares analysis",
    )
    imputed_ls = require(args.imputed_work / "ZI_singer10rooted_LeastSquares_analysis.txt", "imputed least-squares analysis")
    deconv_ls = require(
        args.deconv_work / f"{args.prefix}_deconv0.03_LeastSquares_analysis.txt",
        "deconvolution least-squares analysis",
    )
    expression_dir = args.bugfix_dir / "gene_expression_multinomial_log_expression"
    scaled_dir = args.bugfix_dir / "expression_scaled_mutation_selection_model_bootstrap200"
    codon_slope_file = require(expression_dir / "codon_expression_slopes_and_fitness.tsv", "codon expression slopes")
    codon_fitness_comparison = require(expression_dir / "codon_fitness_comparison.tsv", "codon fitness comparison")
    scaled_summary = require(scaled_dir / "expression_scaled_selection_summary.tsv", "expression-scaled summary")
    scaled_aggregate = require(scaled_dir / "aggregate_observed_vs_predicted_RSCU.tsv", "expression-scaled aggregate table")

    bootstrap_ci = None
    if args.bootstrap_ci.exists():
        bootstrap_ci = parse_bootstrap_ci(args.bootstrap_ci)
    elif not args.allow_missing_ci:
        raise FileNotFoundError(
            f"missing bootstrap CI summary: {args.bootstrap_ci}\n"
            "Use --allow-missing-ci only for a provisional workbook."
        )

    primary_ci, fitted_ci, estg_ci = bootstrap_ci if bootstrap_ci else ({}, {}, {})
    primary = parse_sfratios_summary(main_summary)
    codon_g = parse_ls_codon_values(main_ls)
    estsfs_p0999_g = parse_ls_codon_values(estsfs_p0999_ls)
    imputed_g = parse_ls_codon_values(imputed_ls)
    deconv_g = parse_ls_codon_values(deconv_ls)
    fitted = fitted_pairs_from_codons(list(primary), codon_g)

    tmp_out = args.output_xlsx.with_name(args.output_xlsx.stem + ".tmp" + args.output_xlsx.suffix)
    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(input_xlsx, tmp_out)
    wb = load_workbook(tmp_out)
    codon_order = existing_first_column_order(wb[SHEET5])

    write_rows(wb[SHEET2], parse_sfs(main_sfs))
    write_rows(
        wb[SHEET3],
        rows_with_ci(["CodonPair (from_to)", "2Ns est", "Low95%CI", "Hi95%CI", "CI width"], primary, primary_ci),
    )
    write_rows(
        wb[SHEET4],
        rows_with_ci(["CodonPair (from_to)", "Est", "Low95%CI", "Hi95%CI", "CI width"], fitted, fitted_ci),
    )
    ordered_codon_g = OrderedDict((codon, codon_g[codon]) for codon in codon_order if codon in codon_g)
    write_rows(
        wb[SHEET5],
        rows_with_ci(["Codon", "g estimate", "Low95%CI", "Hi95%CI", "CI width"], ordered_codon_g, estg_ci),
    )
    table6 = [["Codon", "ARG-rooted main pipeline", "est-sfs 5 outgroups p >= 0.999", "ARG-rooted imputed genomes, n=190", "Deconvolution polarization error 3%"]]
    for codon, main_g in ordered_codon_g.items():
        table6.append([codon, main_g, estsfs_p0999_g.get(codon), imputed_g.get(codon), deconv_g.get(codon)])
    write_rows(wb[SHEET6], table6)
    update_table7(wb[SHEET7], codon_g)
    write_rows(wb[SHEET8], parse_ls_model_fit(main_ls))
    write_rows(wb[SHEET10], convert_table_values(tsv_rows(codon_slope_file)))
    write_rows(wb[SHEET11], build_table11_rows(codon_fitness_comparison, codon_slope_file))
    write_rows(wb[SHEET12], build_table12_rows(scaled_summary, scaled_aggregate))

    wb.save(tmp_out)
    tmp_out.replace(args.output_xlsx)
    print(f"wrote {args.output_xlsx}")
    print(f"Table 2 rows: {len(parse_sfs(main_sfs))}")
    print(f"Table 3 codon pairs: {len(primary)}")
    print(f"Table 5 codons: {len(codon_g)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
